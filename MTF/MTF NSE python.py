import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import zipfile
import io
import datetime
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import time
import argparse
import logging
from logging.handlers import RotatingFileHandler
import csv
import os
# Removed: import math # Used for exponential backoff calculation

# --- Configuration ---
START_DATE = datetime.date(2024, 1, 1)
OUTPUT_FILE = 'MTF_Outstanding_Data.xlsx'
SEARCH_PHRASE = "Net scripwise outstanding at the end of the day"
NSE_BASE_URL = "https://nsearchives.nseindia.com/content/equities/mrg_trading_{date}.zip"
DOWNLOAD_TIMEOUT = 60 # Increased timeout to 60 seconds (from 30) for better reliability
MAX_RETRIES = 8
BACKOFF_FACTOR = 10  # base seconds for exponential backoff (BACKOFF_FACTOR * 2**(attempt-1))

# Browser-like headers (helps avoid some server-side throttling/blocks)
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/114.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
    "Referer": "https://www.nseindia.com/",
}

# File to record permanently failed downloads for later resume attempts
FAILED_QUEUE_FILE = 'failed_attempts.csv'
FAILED_ZIPS_DIR = 'failed_zips'

# NSE Trading holidays 2025 (Official NSE Calendar)
# Source: https://www.nseindia.com/resources/exchange-communication-holidays
TRADING_HOLIDAYS_2025 = {
    "2025-02-26",  # Mahashivratri
    "2025-03-14",  # Holi
    "2025-03-31",  # Id-Ul-Fitr (Ramadan Eid)
    "2025-04-10",  # Shri Mahavir Jayanti
    "2025-04-14",  # Dr. Baba Saheb Ambedkar Jayanti
    "2025-04-18",  # Good Friday
    "2025-05-01",  # Maharashtra Day
    "2025-08-15",  # Independence Day / Parsi New Year
    "2025-08-27",  # Shri Ganesh Chaturthi
    "2025-10-02",  # Mahatma Gandhi Jayanti / Dussehra
    "2025-10-21",  # Diwali Laxmi Pujan
    "2025-10-22",  # Balipratipada
    "2025-11-05",  # Prakash Gurpurb Sri Guru Nanak Dev
    "2025-12-25",  # Christmas
}

# Logger (will be configured in main)
logger = logging.getLogger("mtf_nse")

def extract_amount_from_value(amount_raw):
    """Clean and convert value to float. Returns float or original string if conversion fails."""
    if isinstance(amount_raw, str):
        amount_clean = amount_raw.replace(',', '').strip()
        try:
            return float(amount_clean)
        except ValueError:
            return amount_raw
    return amount_raw

def save_failed_entry(process_date, url, reason, attempts):
    """Append a failed download entry to the failed queue CSV."""
    header = ['date', 'url', 'reason', 'attempts', 'last_attempted']
    row = [process_date.strftime('%Y-%m-%d'), url, str(reason), attempts, datetime.date.today().isoformat()]
    write_header = not os.path.exists(FAILED_QUEUE_FILE)
    try:
        with open(FAILED_QUEUE_FILE, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(header)
            writer.writerow(row)
    except Exception as e:
        logger.warning(f"Could not write to failed queue file: {e}")


def save_failed_zip(date_display, zip_bytes, reason=None):
    """Save the raw ZIP bytes to FAILED_ZIPS_DIR for manual inspection."""
    try:
        Path(FAILED_ZIPS_DIR).mkdir(parents=True, exist_ok=True)
        safe_name = date_display.replace(':', '-').replace(' ', '_')
        filename = f"failed_{safe_name}.zip"
        path = Path(FAILED_ZIPS_DIR) / filename
        with open(path, 'wb') as f:
            f.write(zip_bytes)
        logger.info(f"Saved failed ZIP to {path} (reason: {reason})")
        return str(path)
    except Exception as e:
        logger.warning(f"Could not save failed ZIP for {date_display}: {e}")
        return None


def process_zip_response(response, date_display):
    """Given a requests Response containing a ZIP, extract and parse Excel, return extracted_amount or None."""
    try:
        zip_file = zipfile.ZipFile(io.BytesIO(response.content))
    except zipfile.BadZipFile:
        logger.warning(f"Downloaded file is not a valid ZIP file for {date_display}.")
        save_failed_zip(date_display, response.content, reason='bad_zip')
        return None

    data_filename = next((name for name in zip_file.namelist() if name.endswith(('.xls', '.xlsx', '.csv'))), None)
    if not data_filename:
        logger.warning(f"Could not find an Excel/CSV file inside the zip for {date_display}.")
        save_failed_zip(date_display, response.content, reason='no_excel_or_csv_found')
        return None

    try:
        excel_data = zip_file.read(data_filename)
    except Exception as e:
        logger.warning(f"Error reading {data_filename} from ZIP for {date_display}: {e}")
        save_failed_zip(date_display, response.content, reason=f'read_error:{e}')
        return None

    # Parse Excel and search for the phrase
    # If the file is a CSV, parse using pandas.read_csv
    if data_filename.lower().endswith('.csv'):
        try:
            try:
                df = pd.read_csv(io.BytesIO(excel_data), dtype=str, encoding='utf-8')
            except Exception:
                df = pd.read_csv(io.BytesIO(excel_data), dtype=str, encoding='latin1')

            for r_idx, row in df.iterrows():
                for c_idx, val in enumerate(row):
                    if pd.notna(val) and isinstance(val, str) and SEARCH_PHRASE in val:
                        if c_idx + 1 < len(row):
                            amount_raw = row.iloc[c_idx + 1]
                            if pd.notna(amount_raw):
                                return extract_amount_from_value(amount_raw)

            logger.warning(f"Could not find phrase in CSV for {date_display}.")
            save_failed_zip(date_display, response.content, reason='csv_phrase_not_found')
            return None
        except Exception as e:
            logger.warning(f"CSV parsing failed for {date_display}: {e}")
            save_failed_zip(date_display, response.content, reason=f'csv_parse_error:{e}')
            return None

    # First try openpyxl (good for .xlsx and modern parsing)
    try:
        wb = load_workbook(io.BytesIO(excel_data), data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and SEARCH_PHRASE in cell.value:
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    if target_cell.value is not None:
                        return extract_amount_from_value(target_cell.value)
        logger.debug(f"openpyxl: phrase not found in workbook for {date_display}. Trying pandas fallback.")
    except Exception as e:
        logger.debug(f"openpyxl parsing failed for {date_display}: {e}. Trying pandas fallback.")

    # Fallback: try pandas with openpyxl/xlrd engines across all sheets
    try:
        # Try openpyxl engine first
        try:
            sheets = pd.read_excel(io.BytesIO(excel_data), sheet_name=None, engine='openpyxl')
        except Exception:
            sheets = pd.read_excel(io.BytesIO(excel_data), sheet_name=None, engine='xlrd')

        for sheet_name, df in sheets.items():
            for r_idx, row in df.iterrows():
                for c_idx, val in enumerate(row):
                    if pd.notna(val) and isinstance(val, str) and SEARCH_PHRASE in val:
                        if c_idx + 1 < len(row):
                            amount_raw = row.iloc[c_idx + 1]
                            if pd.notna(amount_raw):
                                return extract_amount_from_value(amount_raw)
        logger.warning(f"Could not find the amount for phrase in Excel sheets for {date_display}.")
        save_failed_zip(date_display, response.content, reason='parse_phrase_not_found')
        return None
    except Exception as e:
        logger.warning(f"Pandas-based parsing also failed for {date_display}: {e}")
        save_failed_zip(date_display, response.content, reason=f'pandas_parse_error:{e}')
        return None


def process_failed_queue(sess, resume_limit=None, dry_run=False):
    """Process entries in FAILED_QUEUE_FILE using provided session.
    If resume_limit is provided, limit number of resumed items.
    If dry_run=True, do not write outputs or mutate files (only log attempts).
    """
    if not os.path.exists(FAILED_QUEUE_FILE):
        logger.debug("No failed queue file found to resume.")
        return
    try:
        with open(FAILED_QUEUE_FILE, 'r', newline='', encoding='utf-8') as f:
            reader = list(csv.DictReader(f))
    except Exception as e:
        logger.warning(f"Could not read failed queue: {e}")
        return

    # Clear the file immediately; we'll re-write failed ones after attempts
    try:
        os.remove(FAILED_QUEUE_FILE)
    except Exception:
        pass

    resumed = 0
    for row in reader:
        if resume_limit is not None and resumed >= resume_limit:
            # re-queue remaining rows
            for rem in reader[resumed:]:
                save_failed_entry(datetime.datetime.strptime(rem['date'], '%Y-%m-%d').date(), rem['url'], rem.get('reason', ''), int(rem.get('attempts', 0)))
            break

        try:
            queued_date = datetime.datetime.strptime(row['date'], '%Y-%m-%d').date()
            queued_url = row['url']
        except Exception:
            continue

        logger.info(f"Resuming failed download for {row['date']}: {queued_url}")
        if dry_run:
            logger.info(f"Dry-run: would attempt resume for {row['date']}")
            resumed += 1
            continue

        try:
            resp = sess.get(queued_url, timeout=DOWNLOAD_TIMEOUT, headers=DEFAULT_HEADERS)
            resp.raise_for_status()
            extracted_amount = process_zip_response(resp, row['date'])
            if extracted_amount is not None:
                append_to_excel(OUTPUT_FILE, [[queued_date.strftime('%Y-%m-%d'), extracted_amount]])
                logger.info(f"Resumed and processed {row['date']} successfully.")
            else:
                save_failed_entry(queued_date, queued_url, 'parse_failed', int(row.get('attempts', 0)) + 1)
        except Exception as e:
            logger.warning(f"Resume attempt failed for {row['date']}: {e}")
            save_failed_entry(queued_date, queued_url, str(e), int(row.get('attempts', 0)) + 1)
        resumed += 1


# --- Utility Functions ---

def get_processed_dates(output_file):
    """Reads the existing output file and returns a set of dates already processed."""
    try:
        # Load the existing data using pandas for easy reading
        df = pd.read_excel(output_file, engine='openpyxl')
        
        # Ensure the 'Date' column is converted to date objects for comparison
        # Handles potential date/datetime inconsistencies in the sheet
        if 'Date' in df.columns:
            # Convert to date format 'YYYY-MM-DD' and store as a set
            return set(pd.to_datetime(df['Date']).dt.date)
        return set()
    except FileNotFoundError:
        # If the file doesn't exist, no dates have been processed
        return set()
    except Exception as e:
        print(f"Warning: Could not read existing data from {output_file}. Starting fresh. Error: {e}")
        return set()

def append_to_excel(output_file, new_data):
    """Appends new data (date and amount) to the output Excel file. Handles batch inserts."""
    if not new_data:
        return
    
    new_df = pd.DataFrame(new_data, columns=['Date', 'Outstanding Amount (INR)'])
    
    try:
        if os.path.exists(output_file):
            existing_df = pd.read_excel(output_file, engine='openpyxl')
            combined = pd.concat([existing_df, new_df], ignore_index=True).drop_duplicates(subset=['Date'], keep='first')
            combined.to_excel(output_file, index=False)
        else:
            new_df.to_excel(output_file, index=False, header=True)
            logger.info(f"Created new file: {output_file}")
    except Exception as e:
        logger.exception(f"FATAL: Failed to append data to Excel file: {e}")


def process_date_range():
    """Main function to iterate through dates, download, process, and save data."""
    
    # 1. Setup Date Range and Tracking
    today = datetime.date.today()
    current_date = START_DATE
    dates_to_process = []
    # Determine end date if CLI provided (END_DATE_CLI may be set by __main__)
    end_date = globals().get('END_DATE_CLI', None) or today
    if end_date > today:
        end_date = today

    # Generate the list of dates to check (inclusive of end_date)
    while current_date <= end_date:
        # Skip weekends (Saturday=5, Sunday=6)
        if current_date.weekday() < 5: 
            dates_to_process.append(current_date)
        current_date += datetime.timedelta(days=1)

    # Get the set of dates already processed from the output file
    processed_dates = get_processed_dates(OUTPUT_FILE)
    print(f"Found {len(processed_dates)} dates already processed in {OUTPUT_FILE}.")

    new_data_entries = []
    batch_size = 10  # Batch write every N entries to reduce Excel I/O overhead
    
    # Prepare a requests Session with urllib3 Retry mounted (robust retries)
    session = requests.Session()
    retry_strategy = Retry(
        total=MAX_RETRIES,
        backoff_factor=BACKOFF_FACTOR,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    # Attempt to resume failed downloads (best-effort) before main loop
    process_failed_queue(session)

    # Main per-date processing loop using the session and centralized response processing
    for process_date in dates_to_process:
        # Skip if already processed
        if process_date in processed_dates:
            logger.debug(f"Skipping {process_date.strftime('%Y-%m-%d')}: Already processed.")
            continue

        # Skip if it's a trading holiday
        date_str = process_date.strftime('%Y-%m-%d')
        if date_str in TRADING_HOLIDAYS_2025:
            logger.debug(f"Skipping {date_str}: NSE Holiday")
            continue

        # Download the file from the same day (T instead of T+1)
        download_date = process_date
        
        # Skip if download_date is in future beyond today
        if download_date > today:
            logger.debug(f"Skipping {process_date.strftime('%Y-%m-%d')}: Download date is in future.")
            continue
        
        date_str_url = download_date.strftime('%d%m%y')
        url = NSE_BASE_URL.format(date=date_str_url)
        date_display = process_date.strftime('%Y-%m-%d')

        try:
            logger.info(f"Processing {date_display}: {url}")
            resp = session.get(url, timeout=DOWNLOAD_TIMEOUT, headers=DEFAULT_HEADERS)
            resp.raise_for_status()
        except requests.exceptions.HTTPError as err:
            status = err.response.status_code if hasattr(err, 'response') else None
            if status == 404:
                logger.info(f"File not found (404) for {date_display}. Likely an NSE Holiday.")
                continue
            else:
                logger.warning(f"HTTP Error {status} for {date_display}: {err}")
                save_failed_entry(process_date, url, str(err), MAX_RETRIES)
                continue
        except requests.exceptions.RequestException as err:
            logger.warning(f"Network Error for {date_display}: {err}")
            save_failed_entry(process_date, url, str(err), MAX_RETRIES)
            continue

        extracted_amount = process_zip_response(resp, date_display)
        if extracted_amount is None:
            save_failed_entry(process_date, url, 'parse_failed', MAX_RETRIES)
            continue

        # Batch entries for efficiency
        new_data_entries.append([process_date.strftime('%Y-%m-%d'), extracted_amount])
        if len(new_data_entries) >= batch_size:
            append_to_excel(OUTPUT_FILE, new_data_entries)
            new_data_entries = []

        # Be kind to the NSE server
        time.sleep(0.5)
    
    # Write remaining batched entries
    if new_data_entries:
        append_to_excel(OUTPUT_FILE, new_data_entries)

    print("\n--- Processing Complete ---")
    
    if len(get_processed_dates(OUTPUT_FILE)) > len(processed_dates):
        print(f"Successfully added new data to {OUTPUT_FILE}.")
    else:
        print("No new data was found or processed.")


def configure_logging(log_file=None, level=logging.INFO):
    """Configure module logger and console + rotating file handler."""
    logger.setLevel(level)
    formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')

    # Console handler
    ch = logging.StreamHandler()
    ch.setLevel(level)
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    # File handler
    if log_file:
        fh = RotatingFileHandler(log_file, maxBytes=2_000_000, backupCount=5, encoding='utf-8')
        fh.setLevel(level)
        fh.setFormatter(formatter)
        logger.addHandler(fh)



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MTF outstanding scraper")
    parser.add_argument("--resume-failed", action="store_true", help="Only resume entries from failed_attempts.csv and exit")
    parser.add_argument("--start-date", type=str, help="Override start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", type=str, help="Override end date (YYYY-MM-DD)")
    parser.add_argument("--max-retries", type=int, help="Override MAX_RETRIES for this run")
    parser.add_argument("--backoff-factor", type=int, help="Override BACKOFF_FACTOR for this run")
    parser.add_argument("--log-file", type=str, help="Path to log file (rotating)")
    parser.add_argument("--dry-run", action="store_true", help="Do not write outputs; useful with --resume-failed")
    parser.add_argument("--resume-limit", type=int, help="Limit number of entries to resume from failed queue")

    args = parser.parse_args()

    # Configure logging
    configure_logging(log_file=args.log_file or 'mtf_nse.log')

    # Apply overrides
    if args.max_retries:
        MAX_RETRIES = args.max_retries
        logger.info(f"MAX_RETRIES overridden to {MAX_RETRIES}")
    if args.backoff_factor:
        BACKOFF_FACTOR = args.backoff_factor
        logger.info(f"BACKOFF_FACTOR overridden to {BACKOFF_FACTOR}")
    if args.start_date:
        try:
            START_DATE = datetime.datetime.strptime(args.start_date, '%Y-%m-%d').date()
        except Exception as e:
            logger.error(f"Invalid --start-date: {e}")
            raise
    if args.end_date:
        try:
            END_DATE_CLI = datetime.datetime.strptime(args.end_date, '%Y-%m-%d').date()
        except Exception as e:
            logger.error(f"Invalid --end-date: {e}")
            raise

    # If user only wants to resume failed entries
    if args.resume_failed:
        # Create session with retry strategy
        session = requests.Session()
        retry_strategy = Retry(total=MAX_RETRIES, backoff_factor=BACKOFF_FACTOR, status_forcelist=[429,500,502,503,504], allowed_methods=["HEAD","GET","OPTIONS"])
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        process_failed_queue(session, resume_limit=args.resume_limit, dry_run=args.dry_run)
        logger.info("Resume-only run complete.")
    else:
        process_date_range()